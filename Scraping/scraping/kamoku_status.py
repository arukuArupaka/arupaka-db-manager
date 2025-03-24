import openpyxl
import sys
import os
sys.path.append('C:/arupaka-db-manager/Scraping')
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'scraping.settings')
from django.core.wsgi import get_wsgi_application
application = get_wsgi_application()
from scraping.models import Lecture

def extract_course_data(file_path):
    course_data = {}  # {授業コード: (区分, 選考方法)}
    
    # Excelファイルを開く
    wb = openpyxl.load_workbook(file_path, data_only=True)
    
    # シート1~5を処理
    for sheet_num in range(1, 6):
        sheet_name = f"Sheet{sheet_num}"  
        if sheet_name not in wb.sheetnames:
            continue
        
        sheet = wb[sheet_name]

        for row in sheet.iter_rows(min_row=1, values_only=True):  
            classCode = str(row[5]).strip() if row[5] else None  # F列（授業コード）
            category = str(row[1]).strip() if row[1] else None  # B列（区分）
            selection_method = "抽選" if row[11] and str(row[11]).strip() == "抽選" else "なし"  # L列（選考方法）
            
            if classCode and category:
                # 同じclassCodeの場合、categoryをリストとして追加
                if classCode not in course_data:
                    course_data[classCode] = []
                course_data[classCode].append(category)

    wb.close()
    return course_data

def update_lecture_category(file_path):
    # ハッシュテーブルを作成
    course_data = extract_course_data(file_path)

    # Djangoのデータベースと照合して更新
    lectures = Lecture.objects.filter(classCode__in=course_data.keys())

    for lecture in lectures:
        # classCodeの型を合わせる
        lecture_classCode = str(lecture.classCode)  # ここでclassCodeを文字列に変換

        # course_data.get()でKeyErrorを防ぐ
        categories = course_data.get(lecture_classCode)

        if categories:  
            
            lecture.category = categories[0]
            
            # 変更を保存
            lecture.save()

    print("更新が完了しました")


# 実行
file_path = "./DepInfo/ScienceandEngineering_lecture_data.xlsx"
update_lecture_category(file_path)